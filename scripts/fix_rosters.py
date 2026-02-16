#!/usr/bin/env python3
"""
Parse the league roster Excel file and fix all_team_rosters in poole_team_data.json.

The Excel has a grid layout: 4 teams across in columns (1,3,5,7).
Team headers have a number (like "1", "8 - Plaid Lads", "18 - MACH") in odd columns.
Player names follow in odd columns, emails in even columns.
Spares section starts with "2025 2026 Spares" and should be ignored.
"""
import openpyxl
import json
import copy
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data" / "2025 2026 Curling Roster All Leagues.xlsx"
PTD_PATH = ROOT / "data" / "poole_team_data.json"

wb = openpyxl.load_workbook(XLSX)


def parse_sheet(sheet_name):
    """Parse teams from a roster sheet. Returns dict of {team_name: [players]}."""
    ws = wb[sheet_name]
    team_cols = [1, 3, 5, 7]  # Teams laid out in these columns

    # Find the "2025 2026 Spares" row - everything after is spares, not teams
    spares_row = ws.max_row + 1
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and "2025 2026 Spares" in str(v).strip():
            spares_row = r
            break

    blocks = []  # list of dicts: {name, players}
    active = {c: None for c in team_cols}

    for r in range(1, spares_row):
        for col in team_cols:
            val = ws.cell(r, col).value
            if val is None:
                continue
            val = str(val).strip()
            if not val:
                continue

            email = ws.cell(r, col + 1).value
            has_email = email is not None and '@' in str(email)

            # Skip title rows
            if val in ("Monday Men's Curling League", "Tuesday Men's Curling League",
                       "Team Roster 2025- 2026", "2025/2026 Pools"):
                continue

            # Is this a team header? Starts with digit, no email next to it
            if val[0].isdigit() and not has_email:
                # Save previous block for this column
                if active[col] is not None:
                    blocks.append(active[col])

                # Parse team name from header like "1", "8 - Plaid Lads", "18 - MACH"
                parts = val.split('-', 1)
                team_name = parts[1].strip() if len(parts) > 1 and parts[1].strip() else None
                active[col] = {'name': team_name, 'players': []}

            elif active[col] is not None:
                # Skip non-name entries
                if '@' in val or val.startswith('cell ') or val.startswith('403'):
                    continue
                active[col]['players'].append(val)

    # Flush remaining
    for col in team_cols:
        if active[col] is not None:
            blocks.append(active[col])

    # Build result
    teams = {}
    for b in blocks:
        if not b['players']:
            continue
        name = b['name']
        if name is None:
            # Use first player's last name
            name = b['players'][0].split()[-1]
        teams[name] = b['players']

    return teams


def main():
    monday = parse_sheet("Monday Men's")
    tuesday = parse_sheet("Tuesday Men's")

    # Load current JSON (deep copy for comparison)
    with open(PTD_PATH) as f:
        ptd = json.load(f)
    old_rosters = copy.deepcopy(ptd['all_team_rosters'])

    # Show parsed teams
    for label, data in [("MONDAY", monday), ("TUESDAY", tuesday)]:
        print(f"\n{'='*60}")
        print(f"{label} TEAMS FROM EXCEL ({len(data)} teams)")
        print(f"{'='*60}")
        for name in sorted(data):
            players = data[name]
            print(f"  {name} ({len(players)}): {', '.join(players)}")

    # Build new all_team_rosters - only update Mon/Tue, preserve Friday Mixed
    new_rosters = {}
    all_names = set(monday.keys()) | set(tuesday.keys()) | set(old_rosters.keys())

    for name in sorted(all_names):
        entry = {}
        if name in monday:
            entry["Monday Night"] = monday[name]
        if name in tuesday:
            entry["Tuesday Night"] = tuesday[name]
        # Preserve Friday Night Mixed from old data
        if name in old_rosters and "Friday Night Mixed" in old_rosters[name]:
            entry["Friday Night Mixed"] = old_rosters[name]["Friday Night Mixed"]
        if entry:
            new_rosters[name] = entry

    # Show differences
    print(f"\n{'='*60}")
    print("DIFFERENCES (old -> new)")
    print(f"{'='*60}")

    diffs = 0
    for name in sorted(all_names):
        for league in ["Monday Night", "Tuesday Night"]:
            old_list = old_rosters.get(name, {}).get(league, [])
            new_list = new_rosters.get(name, {}).get(league, [])

            old_set = {p.strip().lower() for p in old_list}
            new_set = {p.strip().lower() for p in new_list}

            if old_set != new_set:
                diffs += 1
                removed = old_set - new_set
                added = new_set - old_set
                print(f"\n  {name} ({league}):")
                for p in sorted(removed):
                    print(f"    REMOVED: {p}")
                for p in sorted(added):
                    print(f"    ADDED:   {p}")

    if diffs == 0:
        print("  No differences found!")
    else:
        print(f"\n  Total roster changes: {diffs}")

    # Update and save
    ptd['all_team_rosters'] = new_rosters
    with open(PTD_PATH, 'w') as f:
        json.dump(ptd, f, indent=2)
    print(f"\n  Updated {PTD_PATH.name}")


if __name__ == '__main__':
    main()
