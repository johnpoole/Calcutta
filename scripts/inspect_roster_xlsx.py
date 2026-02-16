#!/usr/bin/env python3
"""Inspect the league roster Excel to find Ryan Duckworth and MACH."""
import openpyxl

wb = openpyxl.load_workbook('data/2025 2026 Curling Roster All Leagues.xlsx')

for sheet_name in ["Monday Men's", "Tuesday Men's"]:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*60}")
    
    # Print all rows to understand the structure
    for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=False):
        vals = [str(c.value)[:30] if c.value else '' for c in row]
        print(f"  {vals}")
    
    print(f"\n  Total rows: {ws.max_row}, cols: {ws.max_column}")
    
    # Search for Duckworth in all cells
    print(f"\n  Cells containing 'Duckworth' or 'MACH':")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if 'duckworth' in cell.value.lower() or 'mach' in cell.value.lower():
                    print(f"    Row {cell.row}, Col {cell.column}: '{cell.value}'")
