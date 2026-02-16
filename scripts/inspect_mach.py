#!/usr/bin/env python3
"""Show the MACH team roster and Duckworth team roster from the Excel file."""
import openpyxl

wb = openpyxl.load_workbook('data/2025 2026 Curling Roster All Leagues.xlsx')
ws = wb["Tuesday Men's"]

# Show rows around MACH (row 15 and row 37) and Duckworth (row 28 and 46)
print("=== CONTEXT AROUND MACH (row 15, col 11) ===")
for r in range(12, 25):
    vals = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(r, c).value
        if v:
            vals.append(f"C{c}:'{v}'")
    if vals:
        print(f"  Row {r}: {', '.join(vals)}")

print("\n=== CONTEXT AROUND '18 - MACH' (row 37, col 3) - this looks like pool/section ===")
for r in range(34, 52):
    vals = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(r, c).value
        if v:
            vals.append(f"C{c}:'{v}'")
    if vals:
        print(f"  Row {r}: {', '.join(vals)}")

print("\n=== CONTEXT AROUND 'Duckworth' team (row 28, col 11) ===")
for r in range(25, 35):
    vals = []
    for c in range(9, ws.max_column + 1):
        v = ws.cell(r, c).value
        if v:
            vals.append(f"C{c}:'{v}'")
    if vals:
        print(f"  Row {r}: {', '.join(vals)}")

# Also show row 5-6 context for Monday to see Birrell team with Duckworth
ws2 = wb["Monday Men's"]
print("\n=== Monday: Birrell team (team 1) ===")
for r in range(4, 15):
    vals = []
    for c in range(1, 3):
        v = ws2.cell(r, c).value
        if v:
            vals.append(f"C{c}:'{v}'")
    if vals:
        print(f"  Row {r}: {', '.join(vals)}")
