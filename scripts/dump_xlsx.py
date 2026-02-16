#!/usr/bin/env python3
"""
Parse the league roster Excel file and dump all team rosters properly.
Handles the actual grid layout of the spreadsheet.
"""
import openpyxl
import json
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data" / "2025 2026 Curling Roster All Leagues.xlsx"

wb = openpyxl.load_workbook(XLSX)

def dump_sheet(sheet_name):
    ws = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"FULL DUMP: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols)")
    print(f"{'='*80}")
    for r in range(1, ws.max_row + 1):
        vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None:
                s = str(v).strip()
                if s:
                    vals.append(f"C{c}:{repr(s)}")
        if vals:
            print(f"  R{r:3d}: {', '.join(vals)}")

dump_sheet("Monday Men's")
dump_sheet("Tuesday Men's")
